from __future__ import annotations
import os
import zipfile
import shutil
from datetime import datetime, date
from typing import Dict, List, Tuple, Optional, Any

import requests
import pandas as pd

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    HRFlowable, PageBreak, Paragraph, SimpleDocTemplate,
    Spacer, Table, TableStyle, KeepTogether,
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

META_ANUAL = 540
META_MENSAL_ESPERADA = META_ANUAL / 12


MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Marco", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}
MESES_UP  = {k: v.upper() for k, v in MESES_PT.items()}
MESES_ABV = {1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
             7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"}

C_VERM  = colors.HexColor("#CC0000")
C_VERDE = colors.HexColor("#1A7A1A")
C_CAB   = colors.HexColor("#D9D9D9")
C_TOT   = colors.HexColor("#BFBFBF")
C_ALT   = colors.HexColor("#F5F5F5")
C_BRNC  = colors.white
C_PRT   = colors.black


ARCGIS_USER = os.getenv("ARCGIS_USER")
ARCGIS_PASS = os.getenv("ARCGIS_PASS")

TOKEN_URL = (
    "https://observatorio.infraestrutura.mg.gov.br"
    "/portal/sharing/rest/generateToken"
)
DATA_URL = (
    "https://observatorio.infraestrutura.mg.gov.br"
    "/server/rest/services/Hosted"
    "/service_63c7de11c64f4cf3badee2c6d27c850a/FeatureServer/0/query"
)




def _df_vazio() -> pd.DataFrame:
    """Retorna um DataFrame vazio com todas as colunas esperadas pelo sistema."""
    return pd.DataFrame(columns=[
        "data_acao", "data_cadastro", "usuario_responsavel", "usuario",
        "tipo_acao", "tipo_label", "reg_label", "regional", "municipio",
        "n_pessoas_orientadas_educadas", "qtd_material_distribuido",
        "ano", "mes", "dias_atraso", "atrasado",
    ])


def _obter_token() -> str:
    """Obtém token de autenticação ArcGIS."""
    response = requests.post(TOKEN_URL, data={
        "username": ARCGIS_USER,
        "password": ARCGIS_PASS,
        "client":   "referer",
        "referer":  "https://observatorio.infraestrutura.mg.gov.br",
        "f":        "json"
    }, timeout=30)
    data = response.json()
    if "token" not in data:
        raise RuntimeError(f"Falha na autenticacao: {data}")
    return data["token"]


def carregar_dados() -> pd.DataFrame:
    """Carrega dados do ArcGIS. Em caso de falha de conexão ou dados ausentes,
    retorna um DataFrame vazio para que os relatórios sejam gerados sem dados."""
    import logging
    try:
        token = _obter_token()
    except Exception as e:
        logging.warning(f"[carregar_dados] Falha ao obter token ArcGIS: {e}. Retornando dados vazios.")
        return _df_vazio()

    try:
        response = requests.get(DATA_URL, params={
            "where":          "1=1",
            "outFields":      "*",
            "f":              "json",
            "token":          token,
            "returnGeometry": "false"
        }, timeout=60)
        payload = response.json()
    except Exception as e:
        logging.warning(f"[carregar_dados] Falha ao buscar dados ArcGIS: {e}. Retornando dados vazios.")
        return _df_vazio()

    if "features" not in payload or not payload["features"]:
        logging.warning("[carregar_dados] Nenhuma feature retornada pela API. Retornando dados vazios.")
        return _df_vazio()

    df = pd.DataFrame([f["attributes"] for f in payload["features"]])
    return _normalizar_dataframe(df)


def _normalizar_dataframe(df: pd.DataFrame) -> pd.DataFrame:

    df["data_acao"] = pd.to_datetime(df["data_acao"], unit="ms", errors="coerce")
    df = df[df["data_acao"].notna()].copy()

    if "data_cadastro" in df.columns:
        df["data_cadastro"] = pd.to_datetime(df["data_cadastro"], unit="ms", errors="coerce")
    else:
        df["data_cadastro"] = pd.NaT

    df["usuario"] = df.get("usuario_responsavel", None)

    df = df[df["tipo_acao"].notna() & (df["tipo_acao"].astype(str).str.strip() != "")].copy()

    for col in ["n_pessoas_orientadas_educadas", "qtd_material_distribuido"]:
        df[col] = pd.to_numeric(df.get(col, 0), errors="coerce").fillna(0).astype(int)

    COLUNAS_VEICULOS = [
        "numero_veiculos_passeio_blitz",
        "n_veiculos_passageiros_blitz",
        "n_veiculos_carga_blitz",
        "numero_veiculos_moto_blitz",
    ]
    
    for coluna in COLUNAS_VEICULOS:
        df[coluna] = pd.to_numeric(
            df.get(coluna, 0),
            errors="coerce"
        ).fillna(0).astype(int)
    
    df["veiculos"] = df[COLUNAS_VEICULOS].sum(axis=1)

    

    df["tipo_label"] = df["tipo_acao"].apply(_padronizar_tipo_acao)
    df["reg_label"]  = df.apply(_extrair_regional, axis=1)
    df["ano"]        = df["data_acao"].dt.year
    df["mes"]        = df["data_acao"].dt.month

    df["dias_atraso"] = (df["data_cadastro"] - df["data_acao"]).dt.days
    df["atrasado"]    = df["dias_atraso"] > 0
    df["dias_atraso"] = df["dias_atraso"].fillna(0).astype(int)
    df["atrasado"]    = df["atrasado"].fillna(False)

    return df.reset_index(drop=True)


def _padronizar_tipo_acao(valor: str) -> str:
    mapa = {
        "blitz":             "Blitz Educativa",
        "comando_educativo": "Blitz Educativa",
        "curso":             "Curso",
        "palestra":          "Palestra",
        "carreata":          "Carreata",
        "panfletagem":       "Panfletagem",
        "evento":            "Evento Educativo",
        "teatro":            "Teatro Educativo",
        "seminario":         "Seminario",
        "workshop":          "Workshop",
    }
    valor_normalizado = str(valor).lower().strip()
    for chave, label in mapa.items():
        if chave in valor_normalizado:
            return label
    return str(valor).strip().title()


def _extrair_regional(row: pd.Series) -> str:
    regional = str(row.get("regional", "")).strip()
    if regional and regional.lower() not in ("nan", "none", ""):
        partes = regional.split(" - ")
        if len(partes) == 2:
            return f"{partes[0].replace(' URG', '').strip()} - {partes[1].strip()}"
        return regional
    municipio = str(row.get("municipio", "")).strip()
    if municipio and municipio.lower() not in ("nan", "none", ""):
        return municipio
    return "Sede"



def filtrar_por_periodo(df: pd.DataFrame, ano: Optional[int] = None,
                        mes: Optional[int] = None) -> pd.DataFrame:
    """Filtra o DataFrame por ano e/ou mês."""
    df_f = df.copy()
    if ano is not None:
        df_f = df_f[df_f["ano"] == int(ano)]
    if mes is not None:
        df_f = df_f[df_f["mes"] == int(mes)]
    return df_f.reset_index(drop=True)


class DadosAgregados:
   

    def __init__(self, df: pd.DataFrame):
        self.df_original    = df
        self._cache_mensal  = {}
        self._cache_anual   = {}
        self._cache_acumulado = {}

    def get_mensal(self, ano: int, mes: int) -> pd.DataFrame:
        key = (ano, mes)
        if key not in self._cache_mensal:
            self._cache_mensal[key] = filtrar_por_periodo(self.df_original, ano, mes)
        return self._cache_mensal[key]

    def get_anual(self, ano: int) -> pd.DataFrame:
        if ano not in self._cache_anual:
            self._cache_anual[ano] = filtrar_por_periodo(self.df_original, ano)
        return self._cache_anual[ano]

    def get_acumulado(self, ano: int, mes: int) -> pd.DataFrame:
        key = (ano, mes)
        if key not in self._cache_acumulado:
            df_ano = self.get_anual(ano)
            df_ac  = df_ano[df_ano["mes"] <= mes] if not df_ano.empty else df_ano
            self._cache_acumulado[key] = df_ac.reset_index(drop=True)
        return self._cache_acumulado[key]


def _paragrafo(texto: str, tamanho: int = 8, alinhamento: str = "CENTER",
               negrito: bool = False, cor=colors.black) -> Paragraph:
    texto_fmt = "" if (texto is None or str(texto).lower() in ("nan", "none") or texto == 0) else str(texto)
    estilo = ParagraphStyle(
        "custom",
        fontName  = "Helvetica-Bold" if negrito else "Helvetica",
        fontSize  = tamanho,
        alignment = TA_CENTER if alinhamento == "CENTER" else TA_LEFT,
        leading   = tamanho + 2,
        spaceAfter  = 0,
        spaceBefore = 0,
        textColor = cor,
    )
    return Paragraph(texto_fmt, estilo)


def _valor_formatado(valor: Any) -> str:
    if valor and int(valor) != 0:
        return str(int(valor))
    return ""


def _valor_negrito(valor: Any) -> Paragraph:
    return _paragrafo(f"<b>{_valor_formatado(valor)}</b>", tamanho=7.5)


def cabecalho_der_mg(linha3: str) -> Paragraph:
    html = (
        '<para alignment="center"><b>'
        'DEPARTAMENTO DE EDIFICACOES E ESTRADAS DE RODAGEM DO ESTADO DE MINAS GERAIS<br/>'
        'GERENCIA DE EDUCACAO PARA O TRANSITO - DO/GET<br/>'
        f'{linha3}</b></para>'
    )
    return Paragraph(html, ParagraphStyle(
        "header", fontName="Helvetica-Bold", fontSize=10,
        alignment=TA_CENTER, leading=14,
    ))


def frase_educativa(ano: int) -> Paragraph:
    return Paragraph(
        f'No ano de {ano}, a Gerencia de Controle de Infracoes (GCI) esta enviando as autuacoes e '
        f'penalidades cometidas pelos usuarios do transito em todo o Estado de Minas Gerais com a '
        f'seguinte frase educativa: <i>"Desacelere. Seu bem maior e a vida".</i>',
        ParagraphStyle("frase", fontName="Helvetica", fontSize=7,
                       alignment=TA_JUSTIFY, leading=10),
    )


def rodape_pdf() -> list:
    return [
        Spacer(1, 5 * mm),
        HRFlowable(width="100%", thickness=0.5, color=colors.grey),
        Paragraph(
            f"Ultima atualizacao: {date.today().strftime('%d/%m/%Y')}",
            ParagraphStyle("rodape", fontName="Helvetica-Oblique",
                           fontSize=7, alignment=TA_CENTER),
        ),
    ]


def criar_documento_pdf(caminho: str):
    return SimpleDocTemplate(
        caminho,
        pagesize     = landscape(A4),
        leftMargin   = 1.2 * cm,
        rightMargin  = 1.2 * cm,
        topMargin    = 1.2 * cm,
        bottomMargin = 1.5 * cm,
    )



def tabela_acoes_educativas(df_base: pd.DataFrame, ano: int,
                            agrupar_por: str = "reg_label") -> Tuple[Table, Dict]:
    tipos  = sorted(df_base["tipo_label"].dropna().unique()) if not df_base.empty else []
    labels = sorted(df_base[agrupar_por].dropna().unique())  if not df_base.empty else []

    linha1 = [_paragrafo(f"<b>{ano}</b>", tamanho=8, negrito=True)]
    linha2 = [_paragrafo("")]

    # Largura útil do A4 paisagem com margens de 1.2cm = 29.7 - 2.4 = 27.3cm
    LARGURA_UTIL = 27.3 * cm
    LARGURA_LABEL = 4.2 * cm
    n_tipos = len(tipos)
    # Cada tipo tem 3 sub-colunas (Acoes, Pessoas, Divulg.)
    # Distribuímos o espaço restante igualmente entre as sub-colunas
    if n_tipos > 0:
        espaco_total = LARGURA_UTIL - LARGURA_LABEL

        n_subcolunas = len(tipos) * 3 + (1 if "Blitz Educativa" in tipos else 0)
        
        largura = espaco_total / n_subcolunas
        
        col_acoes = largura
        col_pessoas = largura
        col_veiculos = largura
        col_divulg = largura
    else:
        col_acoes = col_pessoas = col_divulg = 2.0 * cm

    larguras_colunas = [LARGURA_LABEL]

    for tipo in tipos:
    
        if tipo == "Blitz Educativa":
    
            linha1 += [
                _paragrafo(f"<b>{tipo.upper()}</b>", tamanho=7.5, negrito=True),
                "", "", ""
            ]
    
            linha2 += [
                _paragrafo("<b>Acoes</b>", tamanho=7, negrito=True),
                _paragrafo("<b>Pessoas</b>", tamanho=7, negrito=True),
                _paragrafo("<b>Veiculos</b>", tamanho=7, negrito=True),
                _paragrafo("<b>Divulg.</b>", tamanho=7, negrito=True),
            ]
    
            larguras_colunas += [
                col_acoes,
                col_pessoas,
                col_veiculos,
                col_divulg
            ]
    
        else:
    
            linha1 += [
                _paragrafo(f"<b>{tipo.upper()}</b>", tamanho=7.5, negrito=True),
                "",
                ""
            ]
    
            linha2 += [
                _paragrafo("<b>Acoes</b>", tamanho=7, negrito=True),
                _paragrafo("<b>Pessoas</b>", tamanho=7, negrito=True),
                _paragrafo("<b>Divulg.</b>", tamanho=7, negrito=True),
            ]
    
            larguras_colunas += [
                col_acoes,
                col_pessoas,
                col_divulg
            ]
    linhas = [linha1, linha2]
    totais = {tipo: {"acoes": 0, "pessoas": 0, "veiculos": 0, "divulg": 0} for tipo in tipos}

    for label in labels:
        subset_label = df_base[df_base[agrupar_por] == label] if not df_base.empty else df_base
        linha = [_paragrafo(label, tamanho=7.5, alinhamento="LEFT")]
        for tipo in tipos:
            subset_tipo = subset_label[subset_label["tipo_label"] == tipo] if not subset_label.empty else subset_label
            acoes  = len(subset_tipo)
            pessoas = int(subset_tipo["n_pessoas_orientadas_educadas"].sum()) if not subset_tipo.empty else 0
            divulg  = int(subset_tipo["qtd_material_distribuido"].sum())      if not subset_tipo.empty else 0
            if tipo == "Blitz Educativa":

                veiculos = int(subset_tipo["veiculos"].sum()) if not subset_tipo.empty else 0
            
                linha += [
                    _valor_formatado(acoes),
                    _valor_formatado(pessoas),
                    _valor_formatado(veiculos),
                    _valor_formatado(divulg)
                ]
            
                totais[tipo]["veiculos"] += veiculos
            
            else:
            
                linha += [
                    _valor_formatado(acoes),
                    _valor_formatado(pessoas),
                    _valor_formatado(divulg)
                ]
            
            totais[tipo]["acoes"] += acoes
            totais[tipo]["pessoas"] += pessoas
            totais[tipo]["divulg"] += divulg
        linhas.append(linha)

    linha_total = [_paragrafo("<b>TOTAL</b>", tamanho=7.5, negrito=True)]
    for tipo in tipos:
    
        t = totais[tipo]
    
        if tipo == "Blitz Educativa":
    
            linha_total += [
                _valor_negrito(t["acoes"]),
                _valor_negrito(t["pessoas"]),
                _valor_negrito(t["veiculos"]),
                _valor_negrito(t["divulg"])
            ]
    
        else:
    
            linha_total += [
                _valor_negrito(t["acoes"]),
                _valor_negrito(t["pessoas"]),
                _valor_negrito(t["divulg"])
            ]
    linhas.append(linha_total)

    estilo = TableStyle([
        ("BOX",        (0, 0), (-1, -1), 1,   C_PRT),
        ("INNERGRID",  (0, 0), (-1, -1), 0.4, C_PRT),
        ("BACKGROUND", (0, 0), (-1,  1), C_CAB),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0,-1), (-1, -1), C_TOT),
        ("FONTNAME",   (0, 0), (-1,  1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
    ])
    col = 1

    for tipo in tipos:
        if tipo == "Blitz Educativa":
            estilo.add("SPAN", (col, 0), (col + 3, 0))
            col += 4
        else:
            estilo.add("SPAN", (col, 0), (col + 2, 0))
            col += 3
    for i in range(2, len(linhas) - 1):
        if i % 2 == 0:
            estilo.add("BACKGROUND", (0, i), (-1, i), C_ALT)

    tabela = Table(linhas, colWidths=larguras_colunas, repeatRows=2)
    tabela.setStyle(estilo)
    return tabela, totais


def bloco_metas(totais: Dict) -> Table:
    total_acoes   = sum(v["acoes"]   for v in totais.values())
    total_pessoas = sum(v["pessoas"] for v in totais.values())
    percentual    = total_acoes / META_ANUAL * 100 if META_ANUAL else 0

    dados = [
        [_paragrafo("<b>TOTALIZADOR GERAL</b>", tamanho=9, negrito=True), "", "",
         _paragrafo(f"<b>META ANUAL: {META_ANUAL}</b>", tamanho=8, negrito=True)],
        [_paragrafo("<b>TIPO DE ACAO</b>", negrito=True),
         _paragrafo("<b>ACOES</b>",        negrito=True),
         _paragrafo("<b>PESSOAS</b>",      negrito=True),
         _paragrafo("<b>% DA META</b>",    negrito=True)],
    ]
    for tipo, v in totais.items():
        perc_tipo = v["acoes"] / META_ANUAL * 100 if META_ANUAL else 0
        dados.append([
            _paragrafo(tipo, tamanho=8),
            _paragrafo(_valor_formatado(v["acoes"]),   tamanho=8),
            _paragrafo(_valor_formatado(v["pessoas"]), tamanho=8),
            _paragrafo(f"{perc_tipo:.1f}%",            tamanho=8),
        ])
    dados.append([
        _paragrafo("<b>TOTAL ACUMULADO</b>", negrito=True),
        _paragrafo(f"<b>{total_acoes}</b>",  negrito=True),
        _paragrafo(f"<b>{total_pessoas}</b>", negrito=True),
        _paragrafo(f"<b>{percentual:.1f}%</b>", negrito=True),
    ])

    tabela = Table(dados, colWidths=[7*cm, 4*cm, 5*cm, 5*cm])
    tabela.setStyle(TableStyle([
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1,  1), C_VERM),
        ("TEXTCOLOR",  (0, 0), (-1,  1), colors.white),
        ("SPAN",       (0, 0), ( 2,  0)),
        ("BACKGROUND", (0,-1), (-1, -1), C_TOT),
    ]))
    return tabela


def grafico_barras_simples(totais: Dict, largura: int = 380, altura: int = 150) -> Drawing:
    categorias = list(totais.keys())
    valores    = [totais[t]["acoes"] for t in categorias]

    desenho   = Drawing(largura, altura + 30)
    bar_chart = VerticalBarChart()
    bar_chart.x      = 60
    bar_chart.y      = 25
    bar_chart.height = altura
    bar_chart.width  = largura - 80
    bar_chart.data   = [valores] if valores else [[0]]
    bar_chart.strokeColor = C_PRT
    bar_chart.categoryAxis.categoryNames  = categorias
    bar_chart.categoryAxis.labels.fontSize = 7
    bar_chart.categoryAxis.labels.angle    = 15
    bar_chart.valueAxis.valueMin  = 0
    bar_chart.valueAxis.valueMax  = max(valores) + max(valores) // 5 if valores else 10
    bar_chart.valueAxis.labels.fontSize = 7
    bar_chart.bars[0].fillColor = C_VERM
    desenho.add(bar_chart)
    return desenho


def construir_story_padrao(titulo: Paragraph, df_base: pd.DataFrame,
                           ano: int, agrupar_por: str = "reg_label") -> list:
    story = [titulo, Spacer(1, 4 * mm)]
    if df_base.empty:
        story.append(_paragrafo("Sem dados para o periodo selecionado.", tamanho=10))
        story.extend(rodape_pdf())
        return story

    tabela, totais = tabela_acoes_educativas(df_base, ano, agrupar_por)
    story.append(tabela)
    story.append(Spacer(1, 5 * mm))
    story.append(frase_educativa(ano))
    story.append(Spacer(1, 5 * mm))

    bloco_meta  = bloco_metas(totais)
    tabela_meta = Table([[bloco_meta]], colWidths=[26.7 * cm])
    tabela_meta.setStyle(TableStyle([("ALIGN", (0, 0), (0, 0), "CENTER")]))
    story.append(tabela_meta)
    story.append(Spacer(1, 8 * mm))

    grafico        = grafico_barras_simples(totais)
    tabela_grafico = Table([[grafico]], colWidths=[26.7 * cm])
    tabela_grafico.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    story.append(KeepTogether([
        _paragrafo("DISTRIBUICAO DE ACOES POR TIPO", tamanho=8, negrito=True),
        Spacer(1, 2 * mm),
        tabela_grafico,
    ]))
    story.extend(rodape_pdf())
    return story




def criar_pasta_mes(base_dir: str, ano: int, mes: int) -> str:
    nome_pasta   = f"{mes:02d}_{MESES_PT[mes]}"
    caminho_pasta = os.path.join(base_dir, str(ano), nome_pasta)
    os.makedirs(caminho_pasta, exist_ok=True)
    return caminho_pasta


def pdf_mensal(dados: DadosAgregados, ano: int, mes: int, diretorio_base: str) -> str:
    df_mensal = dados.get_mensal(ano, mes)
    pasta_mes = criar_pasta_mes(diretorio_base, ano, mes)
    nome_arquivo = f"Relatorio_Mensal_{MESES_PT[mes]}_{ano}.pdf"
    caminho = os.path.join(pasta_mes, nome_arquivo)

    titulo = cabecalho_der_mg(
        f'RELATORIO GERAL DE ACOES EDUCATIVAS — '
        f'<font color="#CC0000">{MESES_UP[mes]} / {ano}</font>'
    )
    criar_documento_pdf(caminho).build(construir_story_padrao(titulo, df_mensal, ano))
    return caminho


def pdf_acumulado(dados: DadosAgregados, ano: int, mes: int, diretorio_base: str) -> str:
    df_acumulado = dados.get_acumulado(ano, mes)
    pasta_mes    = criar_pasta_mes(diretorio_base, ano, mes)
    nome_arquivo = f"Relatorio_Acumulado_{ano}_{mes:02d}.pdf"
    caminho      = os.path.join(pasta_mes, nome_arquivo)

    titulo = cabecalho_der_mg(
        f'RELATORIO ACUMULADO DE ACOES EDUCATIVAS — '
        f'<font color="#CC0000">ACUMULADO {MESES_UP[1]} A {MESES_UP[mes]} / {ano}</font>'
    )
    criar_documento_pdf(caminho).build(construir_story_padrao(titulo, df_acumulado, ano))
    return caminho


def pdf_anual(dados: DadosAgregados, ano: int, diretorio_base: str, pasta_temp: str = None) -> str:
    df_anual = dados.get_anual(ano)
    if pasta_temp:
        os.makedirs(pasta_temp, exist_ok=True)
        caminho = os.path.join(pasta_temp, f"Relatorio_Anual_{ano}.pdf")
    else:
        caminho = os.path.join(diretorio_base, str(ano), f"Relatorio_Anual_{ano}.pdf")
        os.makedirs(os.path.dirname(caminho), exist_ok=True)

    titulo = cabecalho_der_mg(
        f'RELATORIO ANUAL DE ACOES EDUCATIVAS — '
        f'<font color="#CC0000">{ano}</font>'
    )
    criar_documento_pdf(caminho).build(construir_story_padrao(titulo, df_anual, ano))
    return caminho


def pdf_metas(dados: DadosAgregados, ano: int, diretorio_base: str, pasta_temp: str = None) -> str:
    if pasta_temp:
        os.makedirs(pasta_temp, exist_ok=True)
        caminho = os.path.join(pasta_temp, f"Relatorio_Metas_{ano}.pdf")
    else:
        caminho = os.path.join(diretorio_base, str(ano), f"Relatorio_Metas_{ano}.pdf")
        os.makedirs(os.path.dirname(caminho), exist_ok=True)

    documento = criar_documento_pdf(caminho)
    story     = []
    titulo    = cabecalho_der_mg(
        f'ACOMPANHAMENTO DE METAS — '
        f'<font color="#CC0000">{ano} | META: {META_ANUAL} ACOES</font>'
    )
    story.append(titulo)
    story.append(Spacer(1, 5 * mm))

    cabecalho = [
        _paragrafo("<b>MES</b>",            negrito=True),
        _paragrafo("<b>ACOES NO MES</b>",   negrito=True),
        _paragrafo("<b>PESSOAS</b>",         negrito=True),
        _paragrafo("<b>ACUMULADO</b>",       negrito=True),
        _paragrafo("<b>META ESPERADA</b>",   negrito=True),
        _paragrafo("<b>DIFERENCA</b>",       negrito=True),
        _paragrafo("<b>% META</b>",          negrito=True),
    ]
    linhas            = [cabecalho]
    acumulado         = 0
    dados_percentuais = []

    for mes in range(1, 13):
        df_mes        = dados.get_mensal(ano, mes)
        acoes_mes     = len(df_mes)
        pessoas       = int(df_mes["n_pessoas_orientadas_educadas"].sum()) if not df_mes.empty else 0
        acumulado    += acoes_mes
        meta_esp_ac   = META_MENSAL_ESPERADA * mes
        diferenca     = acumulado - meta_esp_ac
        percentual    = acumulado / META_ANUAL * 100

        dados_percentuais.append(round(percentual, 1))

        cor_dif  = C_VERDE if diferenca >= 0 else C_VERM
        cor_perc = C_VERDE if percentual >= 50 else C_VERM

        linhas.append([
            _paragrafo(MESES_PT[mes], tamanho=8),
            _paragrafo(_valor_formatado(acoes_mes) or "-", tamanho=8),
            _paragrafo(_valor_formatado(pessoas)   or "-", tamanho=8),
            _paragrafo(str(acumulado),                     tamanho=8, negrito=(acoes_mes > 0)),
            _paragrafo(f"{int(meta_esp_ac)}",              tamanho=8),
            _paragrafo(f"{'+' if diferenca >= 0 else ''}{int(diferenca)}", tamanho=8, cor=cor_dif),
            _paragrafo(f"{percentual:.1f}%", tamanho=8, negrito=True, cor=cor_perc),
        ])

    total_acoes   = len(dados.get_anual(ano))
    total_pessoas = int(dados.get_anual(ano)["n_pessoas_orientadas_educadas"].sum()) if not dados.get_anual(ano).empty else 0
    linhas.append([
        _paragrafo("<b>TOTAL</b>",                             negrito=True),
        _paragrafo(f"<b>{total_acoes}</b>",                    negrito=True),
        _paragrafo(f"<b>{total_pessoas}</b>",                  negrito=True),
        _paragrafo(f"<b>{total_acoes}</b>",                    negrito=True),
        _paragrafo(f"<b>{META_ANUAL}</b>",                     negrito=True),
        _paragrafo(f"<b>{total_acoes - META_ANUAL:+d}</b>",    negrito=True),
        _paragrafo(f"<b>{total_acoes / META_ANUAL * 100:.1f}%</b>", negrito=True),
    ])

    est = TableStyle([
        ("BOX",        (0, 0), (-1, -1), 1,   C_PRT),
        ("INNERGRID",  (0, 0), (-1, -1), 0.4, C_PRT),
        ("BACKGROUND", (0, 0), (-1,  0), C_VERM),
        ("TEXTCOLOR",  (0, 0), (-1,  0), colors.white),
        ("BACKGROUND", (0,-1), (-1, -1), C_TOT),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    for i in range(1, len(linhas) - 1):
        if i % 2 == 0:
            est.add("BACKGROUND", (0, i), (-1, i), C_ALT)

    tabela = Table(linhas, colWidths=[4*cm, 4*cm, 4.5*cm, 4*cm, 4*cm, 3.5*cm, 4*cm])
    tabela.setStyle(est)
    story.append(tabela)
    story.append(Spacer(1, 8 * mm))

    desenho   = Drawing(420, 170)
    bar_chart = VerticalBarChart()
    bar_chart.x      = 55;  bar_chart.y      = 25
    bar_chart.height = 120; bar_chart.width  = 340
    bar_chart.data   = [dados_percentuais]
    bar_chart.strokeColor = C_PRT
    bar_chart.categoryAxis.categoryNames   = [MESES_ABV[m] for m in range(1, 13)]
    bar_chart.categoryAxis.labels.fontSize = 7
    bar_chart.valueAxis.valueMin           = 0
    bar_chart.valueAxis.valueMax           = max(max(dados_percentuais) + 10, 110) if dados_percentuais else 110
    bar_chart.valueAxis.labels.fontSize    = 7
    bar_chart.valueAxis.labelTextFormat    = "%g%%"
    bar_chart.bars[0].fillColor            = C_VERM
    desenho.add(bar_chart)

    tabela_grafico = Table([[desenho]], colWidths=[26.7 * cm])
    tabela_grafico.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    story.append(KeepTogether([
        _paragrafo("EVOLUCAO DO % DE META ATINGIDA (ACUMULADO MENSAL)", tamanho=8, negrito=True),
        Spacer(1, 3 * mm),
        tabela_grafico,
    ]))
    story.extend(rodape_pdf())
    documento.build(story)
    return caminho


def pdf_regional(dados: DadosAgregados, ano: int, diretorio_base: str, pasta_temp: str = None) -> str:
    if pasta_temp:
        os.makedirs(pasta_temp, exist_ok=True)
        caminho = os.path.join(pasta_temp, f"Relatorio_Regional_{ano}.pdf")
    else:
        caminho = os.path.join(diretorio_base, str(ano), f"Relatorio_Regional_{ano}.pdf")
        os.makedirs(os.path.dirname(caminho), exist_ok=True)

    documento = criar_documento_pdf(caminho)
    story     = []
    primeiro  = True

    for mes in range(1, 13):
        df_mensal = dados.get_mensal(ano, mes)
        if df_mensal.empty:
            continue
        if not primeiro:
            story.append(PageBreak())
        primeiro = False
        titulo = cabecalho_der_mg(
            f'RELATORIO DE ACOES EDUCATIVAS — '
            f'<font color="#CC0000">{MESES_UP[mes]} / {ano}</font>'
        )
        story.extend(construir_story_padrao(titulo, df_mensal, ano))

    if not story:
        story.append(_paragrafo(f"Sem dados para {ano}.", tamanho=10))
        story.extend(rodape_pdf())

    documento.build(story)
    return caminho


def pdf_sem_regional(dados: DadosAgregados, ano: int, diretorio_base: str, pasta_temp: str = None) -> str:
    df_anual = dados.get_anual(ano).copy()
    df_anual["reg_label"] = "TOTAL GERAL"

    if pasta_temp:
        os.makedirs(pasta_temp, exist_ok=True)
        caminho = os.path.join(pasta_temp, f"Relatorio_SemRegional_{ano}.pdf")
    else:
        caminho = os.path.join(diretorio_base, str(ano), f"Relatorio_SemRegional_{ano}.pdf")
        os.makedirs(os.path.dirname(caminho), exist_ok=True)

    titulo = cabecalho_der_mg(
        f'RELATORIO ANUAL SEM REGIONAL — '
        f'<font color="#CC0000">{ano}</font>'
    )
    criar_documento_pdf(caminho).build(construir_story_padrao(titulo, df_anual, ano))
    return caminho


def _estilo_borda(estilo="thin"):
    lado = Side(style=estilo)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _cabecalho_excel(ws, celula, valor, fundo="CC0000", letra="FFFFFF", tamanho=9):
    cel = ws[celula]
    cel.value      = valor
    cel.font       = Font(bold=True, size=tamanho, color=letra, name="Arial")
    cel.fill       = PatternFill("solid", fgColor=fundo)
    cel.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border     = _estilo_borda()


def _celula_excel(ws, celula, valor, negrito=False, fundo="FFFFFF",
                  alinhamento="center", formato=None):
    cel = ws[celula]
    cel.value      = valor
    cel.font       = Font(bold=negrito, size=9, name="Arial")
    cel.fill       = PatternFill("solid", fgColor=fundo)
    cel.alignment  = Alignment(horizontal=alinhamento, vertical="center")
    cel.border     = _estilo_borda()
    if formato:
        cel.number_format = formato


def _aba_mensal_excel(wb: Workbook, dados: DadosAgregados, ano: int, mes: int):
    df_mensal = dados.get_mensal(ano, mes)
    ws = wb.create_sheet("Relatorio Mensal")
    tipos  = sorted(df_mensal["tipo_label"].dropna().unique()) if not df_mensal.empty else []
    labels = sorted(df_mensal["reg_label"].dropna().unique())  if not df_mensal.empty else []

    num_cols    = 1 + len(tipos) * 3
    ultima_col  = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{ultima_col}1")
    cel = ws["A1"]
    cel.value     = f"RELATORIO MENSAL — {MESES_UP.get(mes, mes)}/{ano}"
    cel.font      = Font(bold=True, size=12, color="CC0000", name="Arial")
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    _cabecalho_excel(ws, "A2", "Regional/Municipio", fundo="D9D9D9", letra="000000")

    col = 2
    for tipo in tipos:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        _cabecalho_excel(ws, f"{get_column_letter(col)}2", tipo, fundo="D9D9D9", letra="000000")
        for sub, ci in [("Acoes", col), ("Pessoas", col + 1), ("Divulg.", col + 2)]:
            _cabecalho_excel(ws, f"{get_column_letter(ci)}3", sub, fundo="D9D9D9", letra="000000", tamanho=8)
        col += 3
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    totais = {t: {"a": 0, "p": 0, "d": 0} for t in tipos}
    for idx, label in enumerate(labels):
        linha      = idx + 4
        cor_fundo  = "FFF2F2" if idx % 2 else "FFFFFF"
        _celula_excel(ws, f"A{linha}", label, fundo=cor_fundo, alinhamento="left")
        col = 2
        for tipo in tipos:
            subset = df_mensal[(df_mensal["reg_label"] == label) & (df_mensal["tipo_label"] == tipo)] if not df_mensal.empty else df_mensal
            acoes  = len(subset)
            pessoas = int(subset["n_pessoas_orientadas_educadas"].sum()) if not subset.empty else 0
            divulg  = int(subset["qtd_material_distribuido"].sum())      if not subset.empty else 0
            totais[tipo]["a"] += acoes
            totais[tipo]["p"] += pessoas
            totais[tipo]["d"] += divulg
            for valor in [acoes or None, pessoas or None, divulg or None]:
                _celula_excel(ws, f"{get_column_letter(col)}{linha}", valor, fundo=cor_fundo)
                col += 1

    linha_total = len(labels) + 4
    _celula_excel(ws, f"A{linha_total}", "TOTAL", negrito=True, fundo="BFBFBF")
    col = 2
    for tipo in tipos:
        t = totais[tipo]
        for valor in [t["a"] or None, t["p"] or None, t["d"] or None]:
            _celula_excel(ws, f"{get_column_letter(col)}{linha_total}", valor, negrito=True, fundo="BFBFBF")
            col += 1

    ws.column_dimensions["A"].width = 24
    for ci in range(2, col):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "A4"


def _aba_anual_excel(wb: Workbook, dados: DadosAgregados, ano: int):
    ws = wb.create_sheet("Anual por Mes")
    ws.merge_cells("A1:F1")
    cel = ws["A1"]
    cel.value     = f"RELATORIO ANUAL — {ano}"
    cel.font      = Font(bold=True, size=12, color="CC0000", name="Arial")
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for i, cab in enumerate(["Mes", "Total Acoes", "Pessoas Orientadas", "Divulgacoes", "Meta Esperada", "% da Meta"]):
        _cabecalho_excel(ws, f"{get_column_letter(i + 1)}2", cab)
    ws.row_dimensions[2].height = 28

    acumulado_acoes = 0
    for idx, mes in enumerate(range(1, 13)):
        linha    = idx + 3
        df_mes   = dados.get_mensal(ano, mes)
        acoes    = len(df_mes)
        pessoas  = int(df_mes["n_pessoas_orientadas_educadas"].sum()) if not df_mes.empty else 0
        divulg   = int(df_mes["qtd_material_distribuido"].sum())      if not df_mes.empty else 0
        acumulado_acoes += acoes
        meta_esp  = META_MENSAL_ESPERADA * mes
        percentual = acumulado_acoes / META_ANUAL if META_ANUAL else 0
        cor_fundo  = "FFF2F2" if idx % 2 else "FFFFFF"

        _celula_excel(ws, f"A{linha}", MESES_PT[mes], fundo=cor_fundo, alinhamento="left")
        _celula_excel(ws, f"B{linha}", acoes   or None, fundo=cor_fundo)
        _celula_excel(ws, f"C{linha}", pessoas or None, fundo=cor_fundo)
        _celula_excel(ws, f"D{linha}", divulg  or None, fundo=cor_fundo)
        _celula_excel(ws, f"E{linha}", int(meta_esp),   fundo=cor_fundo)
        _celula_excel(ws, f"F{linha}", percentual if acoes > 0 else None, fundo=cor_fundo, formato="0.0%")

    linha_total = 15
    _celula_excel(ws, f"A{linha_total}", "TOTAL ANUAL", negrito=True, fundo="BFBFBF", alinhamento="left")
    _celula_excel(ws, f"B{linha_total}", f"=SUM(B3:B{linha_total - 1})", negrito=True, fundo="BFBFBF")
    _celula_excel(ws, f"C{linha_total}", f"=SUM(C3:C{linha_total - 1})", negrito=True, fundo="BFBFBF")
    _celula_excel(ws, f"D{linha_total}", f"=SUM(D3:D{linha_total - 1})", negrito=True, fundo="BFBFBF")

    ws.column_dimensions["A"].width = 14
    for ci in range(2, 7):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.freeze_panes = "A3"


def _aba_regional_excel(wb: Workbook, dados: DadosAgregados, ano: int):
    df_anual = dados.get_anual(ano)
    ws = wb.create_sheet("Por Regional")
    ws.merge_cells("A1:N1")
    cel = ws["A1"]
    cel.value     = f"ACOES EDUCATIVAS POR REGIONAL — {ano}"
    cel.font      = Font(bold=True, size=12, color="CC0000", name="Arial")
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for i, cab in enumerate(["Regional"] + [MESES_ABV[m] for m in range(1, 13)] + ["TOTAL"]):
        _cabecalho_excel(ws, f"{get_column_letter(i + 1)}2", cab)
    ws.row_dimensions[2].height = 18

    regionais   = sorted(df_anual["reg_label"].dropna().unique()) if not df_anual.empty else []
    linha_atual = 3

    for metrica_nome, metrica_key in [("N de Acoes", "a"), ("Pessoas Orientadas", "p"), ("Divulgacoes", "d")]:
        ws.merge_cells(f"A{linha_atual}:N{linha_atual}")
        cel = ws[f"A{linha_atual}"]
        cel.value     = metrica_nome
        cel.font      = Font(bold=True, size=10, name="Arial")
        cel.fill      = PatternFill("solid", fgColor="D9D9D9")
        cel.alignment = Alignment(horizontal="left", vertical="center")
        cel.border    = _estilo_borda()
        ws.row_dimensions[linha_atual].height = 16
        linha_atual += 1

        totais_mes = {m: 0 for m in range(1, 13)}
        for idx, regional in enumerate(regionais):
            cor_fundo  = "FFF2F2" if idx % 2 else "FFFFFF"
            df_reg     = df_anual[df_anual["reg_label"] == regional] if not df_anual.empty else df_anual
            _celula_excel(ws, f"A{linha_atual}", regional, fundo=cor_fundo, alinhamento="left")

            total_reg = 0
            for col_idx, mes in enumerate(range(1, 13)):
                df_mes = filtrar_por_periodo(df_reg, mes=mes) if not df_reg.empty else df_reg
                if metrica_key == "a":
                    valor = len(df_mes)
                elif metrica_key == "p":
                    valor = int(df_mes["n_pessoas_orientadas_educadas"].sum()) if not df_mes.empty else 0
                else:
                    valor = int(df_mes["qtd_material_distribuido"].sum())      if not df_mes.empty else 0
                _celula_excel(ws, f"{get_column_letter(col_idx + 2)}{linha_atual}", valor or None, fundo=cor_fundo)
                total_reg    += valor
                totais_mes[mes] += valor

            _celula_excel(ws, f"N{linha_atual}", total_reg or None, negrito=True, fundo=cor_fundo)
            linha_atual += 1

        _celula_excel(ws, f"A{linha_atual}", "TOTAL", negrito=True, fundo="BFBFBF")
        total_geral = 0
        for col_idx, mes in enumerate(range(1, 13)):
            _celula_excel(ws, f"{get_column_letter(col_idx + 2)}{linha_atual}",
                          totais_mes[mes] or None, negrito=True, fundo="BFBFBF")
            total_geral += totais_mes[mes]
        _celula_excel(ws, f"N{linha_atual}", total_geral or None, negrito=True, fundo="BFBFBF")
        linha_atual += 2

    ws.column_dimensions["A"].width = 24
    for ci in range(2, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 8
    ws.freeze_panes = "B3"


def _aba_metas_excel(wb: Workbook, dados: DadosAgregados, ano: int):
    ws = wb.create_sheet("Metas")
    ws.merge_cells("A1:G1")
    cel = ws["A1"]
    cel.value     = f"ACOMPANHAMENTO DE METAS — {ano} | META: {META_ANUAL} ACOES"
    cel.font      = Font(bold=True, size=12, color="CC0000", name="Arial")
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for i, cab in enumerate(["Mes", "Acoes no Mes", "Pessoas", "Acumulado", "Meta Esperada", "Diferenca", "% da Meta"]):
        _cabecalho_excel(ws, f"{get_column_letter(i + 1)}2", cab)
    ws.row_dimensions[2].height = 28

    acumulado = 0
    for idx, mes in enumerate(range(1, 13)):
        linha    = idx + 3
        df_mes   = dados.get_mensal(ano, mes)
        acoes_mes = len(df_mes)
        pessoas   = int(df_mes["n_pessoas_orientadas_educadas"].sum()) if not df_mes.empty else 0
        acumulado += acoes_mes
        meta_esp  = META_MENSAL_ESPERADA * mes
        diferenca = acumulado - meta_esp
        percentual = acumulado / META_ANUAL if META_ANUAL else 0
        cor_fundo  = "FFF2F2" if idx % 2 else "FFFFFF"

        _celula_excel(ws, f"A{linha}", MESES_PT[mes], fundo=cor_fundo, alinhamento="left")
        _celula_excel(ws, f"B{linha}", acoes_mes or None, fundo=cor_fundo)
        _celula_excel(ws, f"C{linha}", pessoas   or None, fundo=cor_fundo)
        _celula_excel(ws, f"D{linha}", acumulado,         fundo=cor_fundo)
        _celula_excel(ws, f"E{linha}", int(meta_esp),     fundo=cor_fundo)
        _celula_excel(ws, f"F{linha}", f"{'+' if diferenca >= 0 else ''}{int(diferenca)}", fundo=cor_fundo)
        _celula_excel(ws, f"G{linha}", percentual, fundo=cor_fundo, formato="0.0%")

    linha_total = 15
    total_acoes = len(dados.get_anual(ano))
    _celula_excel(ws, f"A{linha_total}", "TOTAL ANUAL", negrito=True, fundo="BFBFBF", alinhamento="left")
    _celula_excel(ws, f"B{linha_total}", total_acoes or None, negrito=True, fundo="BFBFBF")
    _celula_excel(ws, f"D{linha_total}", total_acoes or None, negrito=True, fundo="BFBFBF")
    _celula_excel(ws, f"E{linha_total}", META_ANUAL,          negrito=True, fundo="BFBFBF")
    _celula_excel(ws, f"F{linha_total}", total_acoes - META_ANUAL, negrito=True, fundo="BFBFBF")
    _celula_excel(ws, f"G{linha_total}", total_acoes / META_ANUAL if META_ANUAL else 0,
                  negrito=True, fundo="BFBFBF", formato="0.0%")

    ws.column_dimensions["A"].width = 14
    for ci in range(2, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    chart = BarChart()
    chart.type     = "col"
    chart.grouping = "clustered"
    chart.title    = "Acoes Acumuladas vs Meta Esperada"
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = "Mes"
    chart.style  = 10
    chart.width  = 20
    chart.height = 12
    data_ref = Reference(ws, min_col=4, min_row=2, max_row=linha_total - 1)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=linha_total - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "I3")


def gerar_excel_mensal(dados: DadosAgregados, ano: int, mes: int, diretorio_base: str) -> str:

    wb = Workbook()
    wb.remove(wb.active)

    _aba_mensal_excel(wb, dados, ano, mes)
    _aba_anual_excel(wb, dados, ano)
    _aba_regional_excel(wb, dados, ano)
    _aba_metas_excel(wb, dados, ano)

    pasta_mes    = criar_pasta_mes(diretorio_base, ano, mes)
    nome_arquivo = f"Relatorios_DER_MG_{MESES_PT[mes]}_{ano}.xlsx"
    caminho      = os.path.join(pasta_mes, nome_arquivo)
    wb.save(caminho)
    return caminho
